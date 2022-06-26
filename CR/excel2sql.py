import openpyxl
import argparse
import os

parser = argparse.ArgumentParser(description='Generador de comandos SQL a partir de un archivo Excel para una base de datos')
parser.add_argument('-x', '--excel', type=str, nargs='?', default=None,
                    help='nombre de archivo Excel, el programa buscará el primer archivo .xlsx en la carpeta si este valor no es especificado')
parser.add_argument('-o', '--output', type=str, nargs='?', default='datos.sql',
                    help='nombre de script SQL en salida: no olvidar la extensión ".sql", el nombre por defecto es datos.sql')
parser.add_argument('-t', '--tablas', type=str, nargs='?', default=None,
                    help='escribir únicamente la letra inicial de las hojas que se quieren considerar del archivo Excel en un solo string')
inputArgs = parser.parse_args()

file_sql = open(inputArgs.output, 'w', encoding='utf8')
dir_path = os.path.dirname(os.path.realpath(__file__))

# Si no se especifica un archivo Excel en los argumentos, buscamos uno
excel_encontrado = False
if inputArgs.excel == None:
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            if file.endswith('.xlsx'):
                excel = str(root) + '\\' + str(file)
                excel = excel.replace(dir_path + '\\', '')
                excel_encontrado = True
                print('Archivo Excel seleccionado automáticamente: ' + excel)
                break
        if excel_encontrado:
            break
    if not excel_encontrado:
        print('Error: ningún archivo Excel fue encontrado en la actual carpeta.\nAbortando...')
        exit(0)

# Si se especifica un archivo Excel en los argumentos, comprobamos que existe
else:
    if os.path.isfile(str(inputArgs.excel)):
        excel = inputArgs.excel
        print('Archivo Excel seleccionado: ' + excel)
    else:
        print('Error: el archivo ' + inputArgs.excel + ' no fue encontrado.')
        for root, dirs, files in os.walk(dir_path):
            for file in files:
                if file.endswith('.xlsx'):
                    excel = str(root) + '\\' + str(file)
                    excel = excel.replace(dir_path + '\\', '')
                    excel_encontrado = True
                    print('¿Quizás quiso decir ' + excel + ' ?')
        if excel_encontrado == False:
            print('Error: ningún archivo Excel fue encontrado en la actual carpeta.')
        exit(0)

# Comparacion de las hojas del Excel con el argumento --tablas para determinar que leer
workbook = openpyxl.load_workbook(excel)
hojas_nombres = []
if inputArgs.tablas != None:
    for letra in inputArgs.tablas:
        for hoja in workbook:
            if letra == hoja.title[0] and hoja.title not in hojas_nombres:
                hojas_nombres.append(hoja.title)
                break
else:
    hojas_nombres.append('Mujer')
    print('¿Olvidó el argumento -t? Para más informaciónm use el argumento -h o --help')
print('Hojas a leer: ' + str(hojas_nombres))

# Borrar los datos existentes en las tablas
for nombre in hojas_nombres:
    file_sql.write('DELETE FROM site_web_' + nombre.lower() + ';\n')

# Iterar en todas las hojas para extraer los datos
for nombre in hojas_nombres:
    hoja = workbook[nombre]
    id_key = 1
    file_sql.write('\n')

    # Crear query de la forma
    # INSERT INTO <tabla> (id,...,...,...,...,...)  VALUES
    query_insert = 'INSERT INTO site_web_' + nombre.lower() + '(id'
    for row in hoja.iter_rows(min_row=1, max_row=1, min_col=2, max_col=hoja.max_column):
        for cell in row:
            query_insert = query_insert + ',' + str(cell.value)
    query_insert = query_insert + ') VALUES\n'
    file_sql.write(query_insert)
    
    for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=2, max_col=hoja.max_column):
        # Crear query de la forma
        # (<id>,...,...,...,...,...),
        query_value = '\t(' + str(id_key)
        id_key += 1
        for cell in row:
            if cell.value == None:
                # casilla vacia
                query_value = query_value + ',NULL'
            elif cell.data_type == 's':
                # string (agregar comillas)
                query_value = query_value + ',"' + str(cell.value) + '"'
            elif cell.data_type == 'd':
                # fecha (agregar comillas y transformar a formato DD/MM/AAAA)
                fecha = str(cell.value)
                query_value = query_value + ',"' + fecha[8:10] + '/' + fecha[5:7] + '/' + fecha[0:4] + '"'
            elif cell.data_type == 'n' and cell.value > 1500:
                # anho superior a 1500 (agregar comillas)
                # ATENCION: pueden haber problemas si se superan las 1500 escritoras
                query_value = query_value + ',"' + str(cell.value) + '"'
            else:
                query_value = query_value + ',' + str(cell.value)
        if id_key != hoja.max_row:
            query_value = query_value + '),\n'
        else:
            query_value = query_value + ');\n'
        file_sql.write(query_value)

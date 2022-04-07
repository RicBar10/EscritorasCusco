import openpyxl
import argparse
import os

parser = argparse.ArgumentParser(description='Analizador de argumentos de script')
parser.add_argument('--excel', '-x', type=str, nargs='?', help='nombre de archivo excel', default=None)
parser.add_argument('--output', '-o', type=str, nargs='?', help='nombre de script SQL', default='datos.sql')
inputArgs = parser.parse_args()

file_sql = open(inputArgs.output, 'w', encoding='utf8')
dir_path = os.path.dirname(os.path.realpath(__file__))

# Si no se especifica un archivo Excel en los argumentos, buscamos uno
if inputArgs.excel == None:
    excel_encontrado = False
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            if file.endswith('.xlsx'):
                excel = str(root) + '\\' + str(file)
                excel = excel.replace(dir_path + '\\', '')
                excel_encontrado = True
                print('Archivo Excel seleccionado automaticamente: ' + excel)
    if excel_encontrado == False:
        print('Error: ningun archivo Excel fue encontrado en la presente carpeta')
        exit(0)

# Si se especifica un archivo Excel en los argumentos, comprobamos que existe
else:
    if os.path.isfile(str(inputArgs.excel)):
        excel = inputArgs.excel
        print('Archivo Excel seleccionado: ' + excel)
    else:
        print('Error: el archivo ' + inputArgs.excel + ' no fue encontrado')
        excel_encontrado = False
        for root, dirs, files in os.walk(dir_path):
            for file in files:
                if file.endswith('.xlsx'):
                    excel = str(root) + '\\' + str(file)
                    excel = excel.replace(dir_path + '\\', '')
                    excel_encontrado = True
                    print('Quizas quiso decir ' + excel + ' ?')
        if excel_encontrado == False:
            print('Error: ningun archivo Excel fue encontrado en la presente carpeta')
        exit(0)

workbook = openpyxl.load_workbook(excel)
#hojas_nombres = ['Mujer', 'Publicacion', 'Lugar', 'Ejerce']
hojas_nombres = []
for hoja in workbook:
    nombre = hoja.title
    resp = input('Utilizar la hoja ' + nombre + '? (s/n) ')
    i = 0
    while resp != 's' and resp != 'n' and i < 5:
        resp = input('Utilizar la hoja ' + nombre + '? (s/n) ')
        i += 1
    if resp == 's':
        hojas_nombres.append(nombre)
        file_sql.write('DELETE FROM site_web_' + nombre.lower() + ';\n')

# Iterar en todas las hojas para extraer los datos
for nombre in hojas_nombres:
    hoja = workbook[nombre]
    id_key = 1
    file_sql.write('\n')

    # Crear query de la forma
    # INSERT INTO <tabla> (id,...,...,...,...,...)  VALUES
    query_insert = 'INSERT INTO site_web_' + nombre.lower() + '(id'
    for row in hoja.iter_rows(min_row=1, max_row=1, min_col=2, max_col=hoja.max_column):
        for cell in row:
            query_insert = query_insert + ',' + str(cell.value)
    query_insert = query_insert + ') VALUES\n'
    file_sql.write(query_insert)
    
    for row in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=2, max_col=hoja.max_column):
        # Crear query de la forma
        # (<id>,...,...,...,...,...),
        query_value = '\t(' + str(id_key)
        id_key += 1
        for cell in row:
            if cell.value == None:
                # casilla vacia
                query_value = query_value + ',NULL'
            elif cell.data_type == 's':
                # string (agregar comillas)
                query_value = query_value + ',"' + str(cell.value) + '"'
            elif cell.data_type == 'd':
                # fecha (agregar comillas y transformar a formato DD/MM/AAAA)
                fecha = str(cell.value)
                query_value = query_value + ',"' + fecha[8:10] + '/' + fecha[5:7] + '/' + fecha[0:4] + '"'
            elif cell.data_type == 'n' and cell.value > 1500:
                # anho superior a 1500 (agregar comillas)
                # ATENCION: pueden haber problemas si se superan las 1500 escritoras
                query_value = query_value + ',"' + str(cell.value) + '"'
            else:
                query_value = query_value + ',' + str(cell.value)
        if id_key != hoja.max_row:
            query_value = query_value + '),\n'
        else:
            query_value = query_value + ');\n'
        file_sql.write(query_value)
